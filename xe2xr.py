from sys import argv
from turtle import update
from ciscoconfparse import CiscoConfParse
import re
from pprint import pprint
from jinja2 import Template
from ipaddress import IPv4Address, ip_interface
from openpyxl.reader.excel import load_workbook

def XLSXDictReader(f):
    book = load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)
    return (dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1))
    
class xe2xr_template():
    def __init__(self):
        self.l2_trunk_interface = '''
interface {{interface['interface']}}
 {%- if interface['description'] %}
 description {{interface['description']}}
 {%- endif %}
 load-interval 30
!
'''

        self.l2_trunk_template_1 = '''
interface {{interface['interface']}}.{{vlan}} l2transport
 {%- if interface_vlan['description'] %}
 description {{interface_vlan['description']}}
 {%- endif %}
 encapsulation dot1q {{vlan}} exact
 rewrite ingress tag pop {{interface['native']}} symmetric
!
'''
        self.l2_trunk_template_2 = '''
interface {{interface['interface']}}.{{vlan}} l2transport
 {%- if interface['description'] %}
 description {{interface['description']}}
 {%- endif %}
 encapsulation dot1q {{vlan}} exact
 rewrite ingress tag pop {{interface['native']}} symmetric
!
'''
        self.l2vpn_template = '''
l2vpn
 bridge group BVI
  bridge-domain {{vlan}}
   storm-control multicast kbps 400
   storm-control broadcast kbps 400
   {%- for interface in interfaces %}
   interface {{interface}}.{{vlan}}
   !
   {%- endfor %}
   routed interface BVI{{vlan}}
!
'''
        self.l3_interface_template = '''
{%- for l3_interface in l3_interfaces %}
interface {{l3_interface['interface']}}
 {%- if l3_interface['description'] != None %}
 description {{l3_interface['description']}}
 {%- endif %}
 {%- if l3_interface['vlan'] %}
 encapsulation dot1Q {{l3_interface['vlan']}}
 {%- endif %}
 {%- if vrf != None %}
 vrf {{vrf}}
 {%- endif %}
 {%- if l3_interface['ipv4'] != None %}
 ipv4 address {{l3_interface['ipv4']}} {{l3_interface['netmask']}}
 {%- endif %}
 {%- if l3_interface['vlan']%}
 {%- else %}
 load-interval 30
 {%- endif %}
 {%- if l3_interface['arp_timeout'] %}
 arp timeout {{l3_interface['arp_timeout']}}
 {%- endif %}
 {%- if l3_interface['policy_input'] %}
 service-policy input {{l3_interface['policy_input']}}
 {%- endif %}
 {%- if l3_interface['policy_output'] %}
 service-policy output {{l3_interface['policy_output']}}
 {%- endif %}
 {%- if l3_interface['shutdown'] == True %}
 shutdown
 {%- endif %}
!
{%- if l3_interface['hsrp_enable'] %}
router hsrp
 interface {{l3_interface['interface']}}
  address-family ipv4
  {%- if l3_interface['hsrp_version'] != None %}
  hsrp version {{l3_interface['hsrp_version']}}
  {%- endif %}
  hsrp {{l3_interface['hsrp_group']}}
  {%- if l3_interface['hsrp_preempt'] == True %}
  preempt
  {%- endif %}
  {%- if l3_interface['hsrp_priority'] %}
  priority {{l3_interface['hsrp_priority']}}
  {%- endif %}
  address {{l3_interface['hsrp_ip']}}
!
{%- endif %}
{%- endfor %}
'''
        self.static_route = '''
router static
{%- if vrf != None %} 
 vrf {{vrf}}
{%- endif %}
 address-family ipv4 unicast
  {%- for network in networks %}
  {{network['dst_network']}} {{network['netmask']}} {% if network['interface'] %}{{network['interface']}}{% endif %} {{network['gateway']}} {% if network['bfd_interval'] %}bfd fast-detect minimum-interval {{network['bfd_interval']}} multiplier {{network['bfd_multiplier']}}{% endif %} {% if network['tag'] != None %}tag {{network['tag']}}{% endif %} {% if network['description'] != None %}description {{network['description']}}{% endif %}
  {%- endfor %}
!
'''

        self.vrf_route_target = '''
{%- if vrf != None %} 
vrf {{vrf}}
 address-family ipv4 unicast
 {%- if ip_vrf['map'] != None %}
 import map {{ip_vrf['map']}}
 {%- endif %}
  import route-target
  {%- for import in ip_vrf['import-targets'] %}
   {{import}}
  {%- endfor %}
  !
  export route-target
  {%- for export in ip_vrf['export-targets'] %}
   {{export}}
  {%- endfor %}
  !
 !
!
{%- endif %}
'''

        self.bgp_base = '''
router bgp 23700
 nsr
 bgp router-id <<New IP Loopback>>
 bgp graceful-restart restart-time 120
 bgp graceful-restart stalepath-time 360
 bgp graceful-restart
 bgp log neighbor changes detail
 address-family ipv4 unicast
 !
 address-family vpnv4 unicast
 !
 address-family ipv4 mdt
 !
 neighbor-group RR-Routers
  remote-as 23700
  password encrypted 141A020807126929233872
  update-source Loopback0
  address-family vpnv4 unicast
  !
  address-family ipv4 mdt
  !
 !
 neighbor 10.255.10.4
  use neighbor-group RR-Routers
 !
 neighbor 10.255.12.32
  use neighbor-group RR-Routers
 !
 
'''
        self.bgp_vrf_temp = '''
 {%- if vrf != None %} 
 vrf {{vrf}}
  rd {{ip_vrf['rd']}}
  address-family ipv4 unicast
   {%- if ip_vrf['redistribute_conn'] == 1 %}
   redistribute connected
   {%- endif %}
   {%- if ip_vrf['redistribute_static'] == 1 %}
   redistribute static
   {%- endif %}
  !
 !
 {%- endif %}
'''
    
    def xr_static_route(self, vrf=None, networks=[], l3_interfaces=[], bfd=[]):
        static_route = Template(self.static_route)
        network_routes = []
        for network in networks:
            # bfd configuration detection
            bfd_param = {}
            for interface in l3_interfaces:
                ip_net = ip_interface(interface['ipv4']+'/'+interface['netmask'])
                if network['gateway'] in [d['gateway'] for d in bfd] and IPv4Address(network['gateway']) in ip_net.network and interface['bfd_interval']:
                    network['bfd_interval'] = interface['bfd_interval']
                    network['bfd_multiplier'] = interface['bfd_multiplier']
                    continue
            
            network_routes.append(network)
        result = static_route.render(vrf=vrf, networks=network_routes)
        return result
        

    def xr_trunk_interface(self, trunk_interfaces=[], l3_interfaces=[]):
        trunk_interface_1 = Template(self.l2_trunk_template_1)
        trunk_interface_2 = Template(self.l2_trunk_template_2)
        trunk_interface = Template(self.l2_trunk_interface)
        l2vpn = Template(self.l2vpn_template)
        vlans = []
        result = str()
        for interface in trunk_interfaces:
            result += trunk_interface.render(interface=interface)
            for vlan in interface['vlan']:
                vlans.append(vlan)
                interface_vlan = [d for d in l3_interfaces if 'BVI'+vlan == d['interface']]
                if len(interface_vlan) > 0:
                    result += trunk_interface_1.render(interface=interface, vlan=vlan, interface_vlan=interface_vlan[0])
                else:
                    result += trunk_interface_2.render(interface=interface, vlan=vlan)
        
        for vl in set(vlans):
            interfaces_vl = []
            for interface in trunk_interfaces:
                if vl in interface['vlan']:
                    interfaces_vl.append(interface['interface'])
            result += l2vpn.render(vlan=vl, interfaces=interfaces_vl)
        return result
            
    def xr_l3_interface(self, vrf=None, l3_interfaces=[]):
        l3_interface = Template(self.l3_interface_template)
        result = l3_interface.render(vrf=vrf, l3_interfaces=l3_interfaces)
        return result

    def xr_vrf_target(self, vrf=None, vrf_target_int=[]):
        vrf_target = Template(self.vrf_route_target)
        result = vrf_target.render(vrf=vrf, ip_vrf=vrf_target_int)
        return result
    
    def xr_bgp_vrf(self, vrf=None, vrf_target_int=[]):
        bgp_vrf = Template(self.bgp_vrf_temp)
        result = bgp_vrf.render(vrf=vrf, ip_vrf=vrf_target_int)
        return result
    
class xe2xr():
    def __init__(self, file):
        self.ciscoparse = CiscoConfParse(file)
    
    def find_vrf_interface(self):
        vrfes_cmd = self.ciscoparse.find_objects_w_child(parentspec=r'interface', childspec=r'ip vrf forwarding')
        vrfes = []
        for vrf_cmd in vrfes_cmd:
            for vrf in vrf_cmd.children:
                if 'ip vrf forwarding' in vrf.text:
                    vrf_re = re.search('ip vrf forwarding (\S+)', vrf.text)
                    vrfes.append(vrf_re.group(1))
        
        return set(vrfes)

    def find_bfd_static_route(self):
        bfd_static_cmd = self.ciscoparse.find_objects(r'ip route static bfd')
        bfd_static_list = []

        for bfd_static in bfd_static_cmd:
            bfd_static_dict = {
                'interface' : None,
                'gateway' : None
            }
            bfd = re.search('ip route static bfd (Vlan\S+|Giga\S+|TenGiga\S+)?\s?(\d+\.\d+\.\d+\.\d+)?', bfd_static.text)
            if bfd:
                bfd_static_dict['interface'] = bfd.group(1)
                bfd_static_dict['gateway'] = bfd.group(2)
            bfd_static_list.append(bfd_static_dict)
        return bfd_static_list

    def find_l2_interface(self):
        interfaces_cmd = self.ciscoparse.find_objects_w_child(parentspec=r'interface', childspec=r'switchport')
        interfaces = list()
        for interface_cmd in interfaces_cmd:
            interface = {
                'interface' : None,
                'description' : None,
                'mode' : 'access',
                'vlan' : 1
            }
            trunk = {
                'protocol' : 'dot1q',
                'vlan' : 'all',
                'native' : 1
            }
            int_name = re.search('interface (\S+)', interface_cmd.text)
            interface['interface'] = int_name.group(1).replace('Port-channel', 'Bundle-Ether')
            for cmd in interface_cmd.children:
                if 'description' in cmd.text:
                    int_desc = re.search('description (.+)', cmd.text)
                    interface['description'] = int_desc.group(1)
                elif 'switchport mode trunk' in cmd.text:
                    interface['mode'] = 'trunk'
                elif 'switchport trunk encapsulation' in cmd.text:
                    trunk_protocol = re.search('switchport trunk encapsulation (\w+)', cmd.text)
                    trunk['protocol'] = trunk_protocol.group(1)
                elif 'switchport trunk allowed vlan' in cmd.text:
                    vlans = re.findall('\d+', cmd.text)
                    trunk['vlan'] = vlans
                elif 'switchport trunk native vlan' in cmd.text:
                    native = re.search('\d+', cmd.text)
                    trunk['native'] = int(native.group(0))
                    interface['vlan'] = int(native.group(0))
            if interface['mode'] == 'trunk':
                interface.update(trunk)
            interfaces.append(interface)
        return interfaces
  
    def find_l3_interface(self, vrf='default'):
        interfaces = []
        if vrf != 'default':
            interfaces_cmd = self.ciscoparse.find_objects_w_child(parentspec=r'interface', childspec=r'ip vrf forwarding %s\b' % vrf)
        else:
            interfaces_cmd = self.ciscoparse.find_objects_w_child(parentspec=r'interface', childspec=r'ip address')

        for interface_cmd in interfaces_cmd:
            interface = {
                'interface' : None,
                'vrf' : None,
                'description' : None,
                'ipv4' : None,
                'netmask' : None,
                'shutdown' : False
            }
            hsrp = {
                'hsrp_enable' : False,
                'hsrp_group' : None,
                'hsrp_version' : 1,
                'hsrp_preempt' : False,
                'hsrp_priority' : None,
                'hsrp_ip' : None
            }
            int_name = re.search('interface (\S+)', interface_cmd.text)
            interface['interface'] = int_name.group(1).replace('Vlan', 'BVI')
            for cmd in interface_cmd.children:
                if 'description' in cmd.text:
                    desc = re.search('description (.+)', cmd.text)
                    interface['description'] = desc.group(1)
                elif 'ip address' in cmd.text:
                    ip = re.search('ip address (\S+) (\S+)', cmd.text)
                    if ip:
                        interface['ipv4'] = ip.group(1)
                        interface['netmask'] = ip.group(2)
                elif 'arp timeout' in cmd.text:
                    arp_timeout = re.search('arp timeout (\d+)', cmd.text)
                    interface['arp_timeout'] = int(arp_timeout.group(1))
                elif 'shutdown' in cmd.text:
                    interface['shutdown'] = True
                elif 'standby' in cmd.text:
                    hsrp_version = re.search('standby version (2)', cmd.text)
                    hsrp_preempt = re.search('standby \d+ preempt', cmd.text)
                    hsrp_ip = re.search('standby (\d+) ip (\S+)', cmd.text)
                    hsrp_priority = re.search('standby \d+ priority (\d+)', cmd.text)
                    if hsrp_version:
                        hsrp['hsrp_version'] = int(hsrp_version.group(1))
                    elif hsrp_preempt:
                        hsrp['hsrp_preempt'] = True
                    elif hsrp_ip:
                        hsrp['hsrp_group'] = hsrp_ip.group(1)
                        hsrp['hsrp_ip'] = hsrp_ip.group(2)
                        hsrp['hsrp_enable'] = True
                    elif hsrp_priority:
                        hsrp['hsrp_priority'] = int(hsrp_priority.group(1))
                elif 'ip vrf forwarding' in cmd.text:
                    interface_vrf = re.search('ip vrf forwarding (\S+)', cmd.text)
                    interface['vrf'] = interface_vrf.group(1)
                elif 'bfd interval' in cmd.text:
                    bfd = re.search('bfd interval (\d+) min_rx (\d+) multiplier (\d+)', cmd.text)
                    interface['bfd_interval'] = bfd.group(1)
                    interface['bfd_multiplier'] = bfd.group(3)
                elif 'encapsulation dot1Q' in cmd.text:
                    vlan = re.search('encapsulation dot1Q (\d+)', cmd.text)
                    interface['vlan'] = vlan.group(1)
                elif 'service-policy input' in cmd.text:
                    policy_input = re.search('service-policy input (\S+)', cmd.text)
                    interface['policy_input'] = policy_input.group(1)
                elif 'service-policy output' in cmd.text:
                    policy_output = re.search('service-policy output (\S+)', cmd.text)
                    interface['policy_output'] = policy_output.group(1)
            if hsrp['hsrp_enable'] == True:
                interface.update(hsrp)
            interfaces.append(interface)

        return interfaces
    
    def find_static_route(self, vrf='default'):
        if vrf != 'default':
            static_route_cmd = self.ciscoparse.find_objects(r'ip route vrf %s ' % vrf)
        else:
            static_route_cmd = self.ciscoparse.find_objects(r'ip route [^(vrf)]')
        static_routes = list()
        for static_route in static_route_cmd:
            route = {
                'vrf' : None,
                'dst_network' : None,
                'netmask' : None,
                'gateway' : None,
                'interface' : None,
                'description' : None,
                'tag' : None
            }

            vrf = re.search('ip route vrf (\S+)', static_route.text)
            dst_network = re.search('ip route\s?(vrf)?\s?(\S*) (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+) (Vlan\S+|Giga\S+|TenGiga\S+)?\s?(\d+\.\d+\.\d+\.\d+)', static_route.text)
            tag = re.search('ip route\s?(vrf)?\s?(\S*) (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+) (Vlan\S+|Giga\S+|TenGiga\S+)?\s?(\d+\.\d+\.\d+\.\d+) tag (\d+)', static_route.text)
            description = re.search('ip route\s?(vrf)?\s?(\S*) (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+) (Vlan\S+|Giga\S+|TenGiga\S+)?\s?(\d+\.\d+\.\d+\.\d+)\s?(tag)?\s?(\d*)\s?name (\S+)', static_route.text)
            
            if vrf:
                route['vrf'] = vrf.group(1)
            if dst_network:
                route['dst_network'] = dst_network.group(3)
                route['netmask'] = dst_network.group(4)
                if dst_network.group(5):
                    route['interface'] = dst_network.group(5).replace('Vlan', 'BVI')
                route['gateway'] = dst_network.group(6)
            if description:
                route['description'] = description.group(9)[:30]
            if tag:
                route['tag'] = tag.group(7)
            
            static_routes.append(route)
        
        return static_routes

    def find_vrf_target(self, vrf='default'):
        ip_vrfs_cmd = self.ciscoparse.find_objects_w_child(parentspec=r'ip vrf', childspec=r'route-target')
        ip_vrfs = list()
        for each_ip_vrf in ip_vrfs_cmd:
            ip_vrf = {
                'vrf' : None,
                'rd' : None,
                'import-targets' : None,
                'export-targets' : None,
                'map' : None,
                'import_selection' : None,
                'import_limit': None,
                'redistribute_conn' : None,
                'redistribute_static' : None
            }
            import_list = list()
            export_list = list()
            ip_vrf_name = re.search('ip vrf (\S+)', each_ip_vrf.text)
            ip_vrf['vrf'] = ip_vrf_name.group(1)
            for cmd in each_ip_vrf.children:
                if 'route-target import' in cmd.text:
                    ip_vrf_import = re.search('route-target import (.+)', cmd.text)
                    import_list.append(ip_vrf_import.group(1))
                elif 'route-target export' in cmd.text:
                    ip_vrf_export = re.search('route-target export (.+)', cmd.text)
                    export_list.append(ip_vrf_export.group(1))
                elif 'import map' in cmd.text:
                    ip_vrf_map = re.search('import map (\S+)', cmd.text)
                    ip_vrf['map'] = ip_vrf_map.group(1)
                elif 'rd' in cmd.text:
                    ip_vrf_rd = re.search('rd (.+)', cmd.text)
                    ip_vrf['rd'] = ip_vrf_rd.group(1)
            ip_vrf['import-targets'] = import_list
            ip_vrf['export-targets'] = export_list
            ip_vrfs.append(ip_vrf)
        return ip_vrfs
    
    def find_bgp(self, vrfs):
        bgps_cmd = self.ciscoparse.find_objects(r'router bgp')
        bgps = list()
        updated_vrf = list()
        for each_bgp in bgps_cmd:
            bgp = {
                'number' : None,
                'main_neighbor' : None,
                'peer_neighbor' : None
            }
            neighbors = list()
            neighbor_group = {
                'peer-group' : False,
                'name' : None,
                'remote-as' : None,
                'update-source' : None,
                'password-type' : None,
                'password' : None,
            }
            bgp_number = re.search('router bgp (\d+)', each_bgp.text)
            bgp['number'] = bgp_number.group(1)
            for cmd1 in each_bgp.children:
                """ if 'neighbor' in cmd1.text:
                    if 'peer-group' in cmd1.text and neighbor_group['remote-as'] == None:
                        neighbor_name = re.search('neighbor (.+) peer-group', cmd1.text)
                        neighbor_group['name'] = neighbor_name.group(1)
                        neighbor_group['peer-group'] = True
                        bgp['main_neighbor'] = neighbor_name.group(1)
                    elif 'remote-as' in cmd1.text:
                        neighbor_info = re.search('neighbor (.+) remote-as (\d+)', cmd1.text)
                        if neighbor_group['name'] == neighbor_info.group(1) and neighbor_group['peer-group'] == True:
                            neighbor_group['remote-as'] = neighbor_info.group(2)
                        else:
                            neighbors.append(neighbor_group)
                            neighbor_group['name'] = neighbor_info.group(1)
                            neighbor_group['remote-as'] = neighbor_info.group(2)
                            neighbor_group['peer-group'] = False """
                if 'address-family ipv4 vrf' in cmd1.text:
                    address_vrf_name = re.search('address-family ipv4 vrf (\S+)', cmd1.text)
                    for ip_vrf in vrfs:
                        if ip_vrf['vrf'] == address_vrf_name.group(1):
                            current_vrf = ip_vrf
                    for cmd2 in cmd1.children:
                        if 'import path selection all' in cmd2.text:
                            current_vrf['import_selection'] = 1
                        elif 'import path limit 3' in cmd2.text:
                            current_vrf['import_limit'] = 1
                        elif 'redistribute connected' in cmd2.text:
                            current_vrf['redistribute_conn'] = 1
                        elif 'redistribute static' in cmd2.text:
                            current_vrf['redistribute_static'] = 1
                    updated_vrf.append(current_vrf)
            print("let's see")
        return bgps, updated_vrf



def main():
    config = xe2xr(file=argv[1])
    try:
        db = XLSXDictReader(argv[2])
    except:
        db = None
    print(len(argv)) 
    vrfes = config.find_vrf_interface()
    print(vrfes)
    l2_interfaces = config.find_l2_interface()
    print(l2_interfaces)
    bfd_static = config.find_bfd_static_route()
    print(bfd_static)
    vrf_target = config.find_vrf_target()
    print(vrf_target)
    bgpes, vrf_bgp = config.find_bgp(vrf_target)
    print(bgpes)
    print(vrf_bgp)
    xr_template = xe2xr_template()
    #print(bfd_static)

    with open('xr_config_lab.txt','w') as log:
        print('!--------------------------- layer 2 interfaces ----------------------------------')
        log.write('\n!--------------------------- layer 2 interfaces ----------------------------------')
        
        vlan_interfaces = config.find_l3_interface()
        print(vlan_interfaces)
        if db != None:
            l2_updates = list()
            for i in l2_interfaces:
                for d in db:
                    old = d['old_interface']
                    new = d['new_interface']
                    if old in i['interface']:
                        port = i['interface']
                        i['interface'] = port.replace(old,new)
                        #print(i['interface'])
                        break
                l2_updates.append(i)
            l2_interface_template = xr_template.xr_trunk_interface(l2_updates, vlan_interfaces)
            print(l2_interface_template)
            log.write(l2_interface_template)
        else:
            l2_interface_template = xr_template.xr_trunk_interface(l2_interfaces, vlan_interfaces)
            print(l2_interface_template)
            log.write(l2_interface_template)

        if len(vrfes) > 0: #Translate per VRF yang ada
            #l3 interfaces
            for vrf in vrfes:
                try:
                    db = XLSXDictReader(argv[2])
                except:
                    db = None
                print('!----------------------------- vrf %s -------------------------------' % vrf)
                log.write('\n!----------------------------- vrf %s -------------------------------' % vrf)
                #print(db)
                print('!*** l3 interface ***')
                log.write('\n!*** l3 interface ***')
                l3_interfaces = config.find_l3_interface(vrf)
                #print(l3_interfaces)

                if db != None:
                    db = [ j for j in db]
                    l3_updates = list()
                    for i in l3_interfaces:
                        for j in db:
                            old = j['old_interface']
                            new = j['new_interface']
                            #print(i['interface'])
                            #print(old)
                            if old in i['interface']:
                                i['interface'] = i['interface'].replace(old,new)
                                #print(i['interface'])
                                break
                        l3_updates.append(i)
                    #print(l3_updates)
                    l3_interfaces_template = xr_template.xr_l3_interface(vrf=vrf, l3_interfaces=l3_updates)
                    print(l3_interfaces_template)
                    log.write(l3_interfaces_template)
                else:
                    print(l3_interfaces)
                    #print(l3_interfaces)
                    l3_interfaces_template = xr_template.xr_l3_interface(vrf=vrf, l3_interfaces=l3_interfaces)
                    print(l3_interfaces_template)
                    log.write(l3_interfaces_template)

            #static routing
            for vrf in vrfes:
                print('!----------------------------- vrf %s -------------------------------' % vrf)
                log.write('\n!----------------------------- vrf %s -------------------------------' % vrf)
                try:
                    db = XLSXDictReader(argv[2])
                except:
                    db = None
                l3_interfaces = config.find_l3_interface(vrf)
                static_routes = config.find_static_route(vrf)
                if len(static_routes) > 0:
                    #print('*** static route ***')
                    log.write('\n*** static route ***')
                if db != None:
                    db = [ j for j in db ]
                    l3_updates = list()
                    for i in l3_interfaces:
                        for d in db:
                            old = d['old_interface']
                            new = d['new_interface']
                            if old in i['interface']:
                                i['interface'] = i['interface'].replace(old,new)
                                #print(i['interface'])
                                break
                        l3_updates.append(i)
                    static_route_template = xr_template.xr_static_route(vrf=vrf, networks=static_routes, l3_interfaces=l3_updates, bfd=bfd_static)
                    #print(l3_updates)
                    print(static_route_template)
                    log.write(static_route_template)
                else:
                    static_route_template = xr_template.xr_static_route(vrf=vrf, networks=static_routes, l3_interfaces=l3_interfaces, bfd=bfd_static)
                    print(static_route_template)
                    log.write(static_route_template)
        elif len(vrf_target) > 0: #ini hard script banget
            for i in vrf_target:
                vrf_target_template = xr_template.xr_vrf_target(vrf=i['vrf'], vrf_target_int= i)
                print(vrf_target_template)
                log.write(vrf_target_template)
            if len(vrf_bgp) > 0:
                bgp_template = xr_template.bgp_base
                #ketika udah ada IP Loopback baru
                #bgp_template = Template(xr_template.bgp_base)
                #result_bgp = bgp_template.render(ip_loopback = ...)
                #print(result_bgp)
                #log.write(result_bgp)
                print(bgp_template)
                log.write(bgp_template)
                for i in vrf_bgp:
                    vrf_bgp_template = xr_template.xr_bgp_vrf(vrf=i['vrf'], vrf_target_int= i)
                    print(vrf_bgp_template)
                    log.write(vrf_bgp_template)

if __name__ == "__main__":
    main()